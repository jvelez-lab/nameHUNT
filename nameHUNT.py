import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import re

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        file_label.config(text=f"Selected File: {file_path}")
        process_button.config(state=tk.NORMAL)
        return file_path
    else:
        file_label.config(text="No file selected")
        process_button.config(state=tk.DISABLED)
        return None

def is_name(value):
    # Define a simple pattern for names
    # Assuming names are composed of alphabets and may include spaces, hyphens, and apostrophes
    pattern = re.compile(r"^[A-Za-z\s'-]+$")
    return bool(pattern.match(value))

def process_file():
    file_path = file_label.cget("text").replace("Selected File: ", "")
    names = name_entry.get().split()
    if not names or not file_path:
        messagebox.showerror("Error", "Please select a file and enter names.")
        return
    
    try:
        df = pd.read_excel(file_path)
        workbook = load_workbook(file_path)
        sheet = workbook.active

        fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        log_data = {
            "file_name": os.path.basename(file_path),
            "file_path": file_path,
            "highlighted_rows_count": 0,
            "highlighted_values": []
        }

        for idx, row in df.iterrows():
            row_highlighted = False
            for col in df.columns:
                for name in names:
                    cell_value = str(row[col])
                    if pd.notna(row[col]) and name.strip().lower() in cell_value.lower() and is_name(cell_value):
                        sheet.cell(row=idx+2, column=df.columns.get_loc(col)+1).fill = fill
                        if not row_highlighted:
                            log_data["highlighted_rows_count"] += 1
                            log_data["highlighted_values"].append({"row": idx+2, "value": row[col]})
                            row_highlighted = True
        
        log_file_path = os.path.join(os.path.dirname(file_path), "log_file.txt")
        with open(log_file_path, "w") as log_file:
            log_file.write(f"File Name: {log_data['file_name']}\n")
            log_file.write(f"File Path: {log_data['file_path']}\n")
            log_file.write(f"Highlighted Rows Count: {log_data['highlighted_rows_count']}\n")
            log_file.write("Highlighted Values:\n")
            for item in log_data["highlighted_values"]:
                log_file.write(f"Row: {item['row']}, Value: {item['value']}\n")

        workbook.save(file_path)
        os.startfile(log_file_path)
        messagebox.showinfo("Success", "File processed and log file created.")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

root = tk.Tk()
root.title("nameHUNT")

frame = tk.Frame(root)
frame.pack(pady=20, padx=20)

select_button = tk.Button(frame, text="Select Excel File", command=select_file)
select_button.grid(row=0, column=0, padx=5, pady=5)

file_label = tk.Label(frame, text="No file selected")
file_label.grid(row=0, column=1, padx=5, pady=5)

name_label = tk.Label(frame, text="Enter names (separated by space):")
name_label.grid(row=1, column=0, padx=5, pady=5)

name_entry = tk.Entry(frame, width=50)
name_entry.grid(row=1, column=1, padx=5, pady=5)

process_button = tk.Button(frame, text="Let the hunt begin", state=tk.DISABLED, command=process_file)
process_button.grid(row=2, column=0, columnspan=2, pady=10)

root.mainloop()
